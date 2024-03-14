from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import openpyxl
import json
from datetime import datetime
import zipfile
import os


def set_border(style):
    border = Border(left=Side(style=style),
                right=Side(style=style),
                top=Side(style=style),
                bottom=Side(style=style))
    return border

def format_row(sheet, row_number, columns):
    # Define the font, border, and alignment
    font = Font(name='Arial', size=12)

    #set thin border
    border = set_border('thin')

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Create a fill object with yellow color
    yellow_fill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')
    # Set the row height
    sheet.row_dimensions[row_number].height = 100

    # Apply formatting to each cell in the row
    for col in columns:
        cell = sheet[f'{col}{row_number}']
        cell.font = font
        cell.border = border
        cell.alignment = alignment
        # cell.fill = yellow_fill # This style was discarded

def set_cell_properties(sheet, row, column, value, border=None, alignment=None, font=None):
    '''

    Needs to be integrated into inner registry
    Sets the properties of a cell in a given sheet.

    Parameters:
    - sheet: The sheet object where the cell is located.
    - row: The row index of the cell.
    - column: The column index of the cell.
    - value: The value to be set in the cell.
    - border: (optional) The border style to be applied to the cell.
    - alignment: (optional) The alignment style to be applied to the cell.
    - font: (optional) The font style to be applied to the cell.

    Returns:
    - None
    '''

    cell = sheet.cell(row=row, column=column, value=value)
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if font:
        cell.font = font

def hide_sheets(ab, ss):
    for s in ss:
        ab[s].sheet_state = 'hidden'

def find_last_row_in_col(sheet, col_index):
    """
    Find the last non-empty row in a specific column.

    :param sheet: The worksheet object.
    :param col_index: The index of the column to search, starting from 1 for column A.
    :return: The row number of the last non-empty cell, or None if the column is empty.
    """
    # Openpyxl is 1-indexed, but using max_row directly as start makes the code clearer
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=col_index).value:
            return row
    return None

def load_excel(f):
    ab1 = openpyxl.load_workbook(f)
    return ab1

def read_json():
    '''
    Returns sorted json file from the model data
    Component of the testing module
    '''

    with open('v2/tests/model.json', 'r') as file:
        json_data = json.load(file)

        payment_documents = json_data.get('request', [])

        sorted_payment_documents = sorted(payment_documents, key=lambda x: x.get('object_name', ''))

        json_data = {'request': sorted_payment_documents}
        return json_data

def format_datetime(datetime_str):
    # Parse the datetime string
    try:
        dt = datetime.fromisoformat(datetime_str)

        # Format the datetime into "dd/mm/yyyy" format
        formatted_date = dt.strftime("%d/%m/%Y")
        return formatted_date
    except ValueError:
        return datetime_str


def create_concatenated_info(data_item):
    parts = []

    payment_type = data_item.get('payment_type', '') # Retrieve payment type
    if payment_type:
        parts.append(payment_type)

    payment_objective = data_item.get('payment_objective', '') #Retrieve payment objective (Naznachenie)
    if payment_objective:
        parts.append(payment_objective)

    schet_na_oplatu = data_item.get('schet_na_oplatu', '') # and etc
    if schet_na_oplatu:
        parts.append(f"Счет на оплату №{schet_na_oplatu}")

    esf = data_item.get('esf', '')
    if str(esf).strip("№"):
        parts.append(f"ЭСФ №{esf}")

    avr = data_item.get('avr', '') # Retrieve AVR
    if avr:
        parts.append(f"Акт выполненных работ №{avr}")

    akt_sverki = data_item.get('akt_sverki', '') # Retrieve Akt sverki
    if akt_sverki:
        parts.append(f"Акт сверки №{akt_sverki}")
    sz = data_item.get('sluzhebnaja_zapiska', '')
    if sz:
        parts.append(f'Служебная записка {sz}')

    avansovy_otchet = data_item.get('avansovy_otchet', '') # Retrieve Avansovy Otchet
    if avansovy_otchet:
        parts.append(f"Авансовый отчет №{avansovy_otchet}")

    tru = data_item.get('TRU', '')
    if tru:
        parts.append(tru)

    letter = data_item.get('letter', '')
    if letter:
        parts.append(letter)

    mediation = data_item.get('mediation', '')
    if mediation:
        parts.append(f"Медиация/Решение суда №{mediation}")

    nakladnye = data_item.get('nakladnye', '')
    if nakladnye:
        parts.append(f"Накладные: {nakladnye}")

    prilozhenije = data_item.get('prilozhenija', '')
    if prilozhenije.lstrip('Приложение '):
        parts.append(f'по приложению {prilozhenije}')

    zusaetzliches_vertrag = data_item.get('zusaetzliches_vertrag', '')
    if zusaetzliches_vertrag != 'placeholder' and zusaetzliches_vertrag:
        zv_text = f'{zusaetzliches_vertrag}'.lstrip('Доп. соглашение')
        parts.append(zv_text)
    else:
        name_of_contract = data_item.get('name_of_contract', '')
        date_of_contract = data_item.get('date_of_contract', '') # to be deleted

        if name_of_contract and date_of_contract:
            formatted_date = format_datetime(date_of_contract) # to be deleted
            parts.append(f"Дог. №{name_of_contract}")

    # Join all parts with ", " and remove trailing comma and space if any
    concatenated_info = ", ".join(parts).rstrip(", ")

    return concatenated_info

def add_colontituls(sheet):
    sheet.oddFooter.left.text = "Группа компаний «Шар Құрылыс»" # Left footer
    sheet.oddFooter.right.text = "Дата и время печати &D &T" # Right footer

def set_print_area(sheet):
    last_row = find_last_row_in_col(sheet,6)
    sheet.print_area = f'F1:I{last_row}'

def split_workbook(filename):
    '''
    This function splits the workbook into separate files and then zips them
    into a single file. It also deletes the files after they are added to the zip

    :param filename: The name of the file to be split
    '''

    initial_sheets = ['REESTR', 'СПР_ОБЪЕКТОВ', 'СПР_ПОДПИСАНТОВ']

    workbook = load_excel(f'saves/{filename}') # load a workbook
    filenames = [] # for the future filenames
    date_str = datetime.now().strftime('%d/%m/%Y_%H%M%S') # Format the current date

    for sheet in workbook.sheetnames:
        # if the sheet name is not in initial sheets
        if sheet not in initial_sheets:
            new_workbook = load_excel('template.xlsx')
            new_workbook.remove(new_workbook.active)
            new_workbook.add_sheet(workbook[sheet])
            new_workbook.save(f'saves/{sheet}_{date_str}.xlsx')
            filenames.append(f'saves/{sheet}_{date_str}.xlsx')

    filename.rstrip('.xlsx') # Remove .xlsx from the filename
    zipname = f'zips/{filename}.zip'
    with zipfile.ZipFile(zipname, 'a') as zipf:
        for file in filenames:
            zipf.write(file)
    
    # Delete the files after they are added to the zip
    for file in filenames:
        os.remove(file)

    return zipname
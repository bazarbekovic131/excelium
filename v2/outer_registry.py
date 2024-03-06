'''
pip install Flask, flask-cors, openpyxl, Flask-Limiter

pip install gunicorn

'''
import logging
from scripts import find_last_row_in_col
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from scripts import load_excel, format_row, format_datetime

def add_coordinators_outer(sheet):
    col_index = 2 # Starting from column B

    final_row = find_last_row_in_col(sheet, col_index)
    if final_row:
        logging.info(f"The last non-empty cell in column {chr(64 + col_index)} of sheet '{sheet.title}' is in row {final_row}.")
    else:
        logging.info(f"No non-empty cells found in column {chr(64 + col_index)} of sheet '{sheet.title}'.")

    final_row +=4
    sheet.cell(row=final_row, column=2, value="Уполномоченная компания: ТОО \"New Line Project\"")
    sheet.cell(row=final_row, column=2).alignment = Alignment(horizontal='left')
    sheet.cell(row=final_row, column=2).font = Font(size=14, bold=True)

    sheet.cell(row=final_row, column=6, value = "М.П.") # F-column, 1st row

    #  Writing the director of the company:

    final_row +=2
    sheet.cell(row=final_row, column=2, value="Директор:")
    sheet.cell(row=final_row, column=2).alignment = Alignment(horizontal='left')
    sheet.cell(row=final_row, column=2).font = Font(size=14, bold=True)

    sheet.cell(row=final_row, column=5).border = Border(bottom=Side(style='thick'))
    sheet.cell(row=final_row+1, column=5, value="(подпись)")
    sheet.cell(row=final_row, column=6, value="Бектемирова Ж.Ж")
    sheet.cell(row=final_row, column=6).font = Font(bold=True)

    # Writing the Ispolnitel

    final_row +=3
    sheet.cell(row=final_row, column=2, value="Исполнитель:")
    sheet.cell(row=final_row, column=2).alignment = Alignment(horizontal='left')
    sheet.cell(row=final_row, column=2).font = Font(size=14, bold=True)

    sheet.cell(row=final_row, column=5).border = Border(bottom=Side(style='thick'))
    sheet.cell(row=final_row+1, column=5, value="(подпись)")
    sheet.cell(row=final_row, column=6, value="Олжабаева Г.Т.")
    sheet.cell(row=final_row, column=6).font = Font(bold=True)
    for col in range(2,8):
        sheet.cell(row=final_row+2, column=col).border = Border(bottom = Side(style='thick'))

    # Writing the second company

    final_row +=4
    sheet.cell(row=final_row, column=2, value="ТОО \"Инжиниринговая компания \"Лидер\"\"")
    sheet.cell(row=final_row, column=2).alignment = Alignment(wrap_text=False)

    final_row +=2
    sheet.cell(row=final_row, column=2, value="Исполнитель:")
    sheet.cell(row=final_row, column=5).border = Border(bottom=Side(style='thick'))
    sheet.cell(row=final_row+1, column=5, value="(подпись)")
    sheet.cell(row=final_row, column=6, value="Колоскова И.Б.")

def loop_json_outer(json_data, workbook):
    '''
    This function works with the loaded json and with the copied workbook
    The workbook is the outer registry file
    '''
    columns = cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

    banks = {
    'KCJBKZKX': 'Банк Центр Кредит',
    'KSNVKZKA': 'Банк Фридом Финанс Казахстан',
    '046577743': '?',
    'KKMFKZ2A': '?',
    'KPSTKZKA': 'АО "Казпочта"',
    'KZIBKZKA': 'ДБ "КЗИ Банк"',
    'KINCKZKA': 'Bank RBK',
    'SABRKZKA': '?',
    'IRTYKZKA': 'ForteBank',
    '044525142': '?',
    '044525700': '?',
    'CASPKZKA': 'Kaspi Bank',
    'HSBKKZKX': 'Народный сберегательный банк Казахстана',
    'TSESKZKA': 'First Heartland Jusan Bank',
    'EURIKZKA': 'Евразийский Банк',
    'NBRKKZKX': '?'
    }

    start_row = 13 # starting row for the writing
    for key_title, data in json_data.items():

        logging.info('Fetched len(data)') #how many documents are fetched

        
        for i in range(len(data)):
            row = start_row
            source_sheet = workbook.active

            #objektname speichern
            source_sheet[f'B7'] = f'Разрешение № от {datetime.today()}'

            #Zeile formatieren
            format_row(source_sheet, row, columns)

            osnovanije_str = '' # String that contains data on the osnovanije of the payment for the outer registry

            # On dr JSON datei bekommn
            source_sheet[f'B{row}'] = i + 1 # Nomer poziciji
            source_sheet[f'C{row}'] = data[i]['organization'] # OrgName
            if data[i]['payment_type']:# Gebuhrtyp
                source_sheet[f'E{row}'] = data[i]['payment_type']

            
            if data[i]['zusaetzliches_vertrag']:
                osnovanije_str += f'Доп. соглашение №{data[i]["zusaetzliches_vertrag"]} от {data[i]["date_of_zv"],};' # if there is a chosen additional contract (DS)

            if data[i]['prilozhenija']: ### CHECK NAME PLACEHOLDER AKHMET
                osnovanije_str += f'{data[i]["prilozhenija"]},'
            
            if data[i]['name_of_contract'] and data[i]['date_of_contract']:
                contract_str = f'{data[i]["name_of_contract"]} от {format_datetime(data[i]["date_of_contract"])}'
                osnovanije_str += contract_str

                old_value = str(source_sheet[f'E{row}'].value)
                old_value += f' к договору {contract_str}'
                source_sheet[f'E{row}'] = old_value

                source_sheet[f'F{row}'] = contract_str #VertragName und Datum

            source_sheet[f'G{row}'] = float(data[i]['contract_sum'])

            source_sheet[f'H{row}'] = float(data[i]['payment_sum']) # Gebuhr

            source_sheet[f'D{row}'] = osnovanije_str

            #source_sheet[f'K{row}'] = data[i]['zatraty'] # Kosten

            counter_str = f'ИИК {data[i]["schet_counter"]}' + '\n' + f'БИК {data[i]["BIK_counter"]} в {banks[data[i]["BIK_counter"]]}' + '\n'
            if data[i]["schet_counter"] and data[i]["BIK_counter"]:
                source_sheet[f'I{row}'] = counter_str
            start_row += 1

def format_excel_outer(json_data):

    logging.info('Opening template.xlsx')
    workbook = load_excel('template_outer.xlsx')
    logging.info('Reading JSON file')

    logging.info('Looping through JSON. Adding documents')
    loop_json_outer(json_data, workbook)

    add_coordinators_outer(workbook.active)

    return workbook